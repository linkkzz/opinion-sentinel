import hashlib
import io
import re
import zipfile
from datetime import datetime
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import MediaAsset, SourceItem


HEADERS = {
    "平台": "platform", "platform": "platform",
    "原始数据ID": "external_id", "external_id": "external_id",
    "标题": "title", "title": "title",
    "发布人": "author", "author": "author",
    "发布时间": "publish_time", "publish_time": "publish_time",
    "正文": "content", "content": "content",
    "源链接": "source_url", "source_url": "source_url",
    "点赞量": "like_count", "点赞数": "like_count", "like_count": "like_count",
    "评论量": "comment_count", "评论数": "comment_count", "comment_count": "comment_count",
    "转发量": "share_count", "转发数": "share_count", "分享量": "share_count", "share_count": "share_count",
    "阅读/播放量": "view_count", "阅读量": "view_count", "播放量": "view_count", "view_count": "view_count",
    "互动量": "interaction_count", "interaction_count": "interaction_count",
    "图片文件": "image_files", "image_files": "image_files",
    "视频文件": "video_file", "video_file": "video_file",
}
ALLOWED_MEDIA = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4", ".mov", ".webm"}


def _clean_name(name: str) -> str:
    return Path(name).name


def _parse_datetime(value) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法识别发布时间：{text}")


def _dedupe_key(row: dict) -> str:
    raw = row.get("external_id") or row.get("source_url") or "|".join(
        str(row.get(key, "")) for key in ("platform", "title", "author", "publish_time", "content")
    )
    return hashlib.sha256(str(raw).strip().encode("utf-8")).hexdigest()


def _parse_count(value, field: str) -> int:
    if value in (None, ""):
        return 0
    try:
        count = int(float(str(value).replace(",", "").strip()))
    except ValueError as exc:
        raise ValueError(f"{field}必须是数字") from exc
    if count < 0:
        raise ValueError(f"{field}不能小于0")
    return count


async def import_excel(db: Session, task_id: int, excel: UploadFile, media_zip: UploadFile | None) -> dict:
    workbook = load_workbook(io.BytesIO(await excel.read()), read_only=True, data_only=True)
    sheet = workbook.active
    raw_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    headers = [HEADERS.get(value, value) for value in raw_headers]
    has_split_engagement = any(name in headers for name in ("like_count", "comment_count", "share_count"))
    required = {"platform", "content"}
    if not required.issubset(headers):
        raise ValueError("Excel必须包含“平台”和“正文”列")

    media_files: dict[str, bytes] = {}
    if media_zip:
        with zipfile.ZipFile(io.BytesIO(await media_zip.read())) as archive:
            for info in archive.infolist():
                name = _clean_name(info.filename)
                if not name or Path(name).suffix.lower() not in ALLOWED_MEDIA or info.file_size > 100 * 1024 * 1024:
                    continue
                media_files[name] = archive.read(info)

    task_dir = settings.storage_root / "tasks" / str(task_id)
    task_dir.mkdir(parents=True, exist_ok=True)
    imported = skipped = 0
    errors: list[str] = []
    for number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, values))
        if not any(value not in (None, "") for value in values):
            continue
        try:
            content = str(row.get("content") or "").strip()
            platform = str(row.get("platform") or "").strip()
            if not content or not platform:
                raise ValueError("平台或正文为空")
            like_count = _parse_count(row.get("like_count"), "点赞量")
            comment_count = _parse_count(row.get("comment_count"), "评论量")
            share_count = _parse_count(row.get("share_count"), "转发量")
            view_count = _parse_count(row.get("view_count"), "阅读/播放量")
            interaction_count = (
                like_count + comment_count + share_count
                if has_split_engagement
                else _parse_count(row.get("interaction_count"), "互动量")
            )
            item = SourceItem(
                task_id=task_id,
                platform=platform,
                external_id=str(row["external_id"]).strip() if row.get("external_id") else None,
                title=str(row.get("title") or "无标题").strip(),
                author=str(row.get("author") or "未知").strip(),
                publish_time=_parse_datetime(row.get("publish_time")),
                content=content,
                source_url=str(row["source_url"]).strip() if row.get("source_url") else None,
                like_count=like_count,
                comment_count=comment_count,
                share_count=share_count,
                view_count=view_count,
                interaction_count=interaction_count,
                dedupe_key=_dedupe_key(row),
            )
            db.add(item)
            db.flush()
            names = re.split(r"[,，;；]", str(row.get("image_files") or ""))
            if row.get("video_file"):
                names.append(str(row["video_file"]))
            for raw_name in filter(None, (value.strip() for value in names)):
                name = _clean_name(raw_name)
                if name not in media_files:
                    continue
                stored_name = f"{item.id}_{name}"
                path = task_dir / stored_name
                path.write_bytes(media_files[name])
                media_type = "video" if path.suffix.lower() in {".mp4", ".mov", ".webm"} else "image"
                db.add(MediaAsset(item_id=item.id, media_type=media_type, original_name=name, storage_path=str(path)))
            db.commit()
            imported += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
        except Exception as exc:
            db.rollback()
            errors.append(f"第{number}行：{exc}")
    return {"imported": imported, "skipped": skipped, "errors": errors[:20]}
