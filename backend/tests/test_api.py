import io
import os
from pathlib import Path

os.environ["DATABASE_URL"] = "sqlite:///./test_opinion_sentinel.db"
os.environ["STORAGE_ROOT"] = "./test_storage"

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.database import Base, SessionLocal, engine
from app.models import ReportVersion, StrategyVersion
from app.main import app


def excel_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["平台", "标题", "发布人", "发布时间", "正文", "源链接", "点赞量", "评论量", "转发量", "阅读/播放量"])
    sheet.append(["微博", "校园食堂事件", "校园观察", "2026-06-20 10:00:00", "学生反映食堂卫生问题，校方正在核实。", "https://example.com/1", 50, 28, 10, 2600])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def legacy_excel_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["平台", "正文", "源链接", "互动量"])
    sheet.append(["小红书", "旧模板导入兼容测试。", "https://example.com/legacy", 77])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def setup_function():
    Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)


def test_task_import_and_manual_analysis():
    with TestClient(app) as client:
        response = client.post("/api/tasks", json={
            "name": "校园食品安全",
            "keywords": ["食堂", "食品安全"],
            "platforms": ["微博", "小红书"],
        })
        assert response.status_code == 201
        task_id = response.json()["id"]

        response = client.post(
            f"/api/tasks/{task_id}/import",
            files={"excel": ("data.xlsx", excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 1

        item = client.get(f"/api/tasks/{task_id}/items").json()["items"][0]
        assert item["like_count"] == 50
        assert item["comment_count"] == 28
        assert item["share_count"] == 10
        assert item["interaction_count"] == 88
        assert item["view_count"] == 2600
        response = client.put(f"/api/items/{item['id']}/analysis", json={
            "sentiment": "negative",
            "risk_level": "medium",
            "reason": "涉及食品卫生，但目前传播范围有限。",
            "topics": ["食品安全"],
            "change_note": "人工复核样例",
        })
        assert response.status_code == 200
        assert response.json()["source"] == "human"

        eligibility = client.get(f"/api/tasks/{task_id}/strategies/eligibility").json()
        assert eligibility["state"] == "available"
        assert eligibility["eligible"] is True
        assert eligibility["reason"] == "存在新的研判结果"
        assert eligibility["analyzed_count"] == 1

        stats = client.get(f"/api/tasks/{task_id}/stats").json()
        assert stats["total"] == 1
        assert stats["risks"]["medium"] == 1
        assert stats["engagement"] == {
            "likes": 50, "comments": 28, "shares": 10, "interactions": 88, "views": 2600,
        }
        overview = client.get("/api/tasks/overview/dashboard")
        assert overview.status_code == 200
        assert overview.json()["tasks"] == 1
        assert overview.json()["analysis_rate"] == 100.0
        assert overview.json()["interactions"] == 88
        assert overview.json()["views"] == 2600


def test_legacy_interaction_column_is_still_supported():
    with TestClient(app) as client:
        task_id = client.post("/api/tasks", json={
            "name": "旧模板兼容", "keywords": ["测试"], "platforms": ["小红书"]
        }).json()["id"]
        response = client.post(
            f"/api/tasks/{task_id}/import",
            files={"excel": ("legacy.xlsx", legacy_excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        item = client.get(f"/api/tasks/{task_id}/items").json()["items"][0]
        assert item["interaction_count"] == 77
        assert item["like_count"] == item["comment_count"] == item["share_count"] == 0


def test_updated_sample_excel_imports_split_metrics():
    sample = Path("test_data/舆情导入测试包/01_校园食品安全_第一批20条.xlsx")
    with TestClient(app) as client, sample.open("rb") as excel:
        task_id = client.post("/api/tasks", json={
            "name": "测试包校验", "keywords": ["食堂"], "platforms": ["微博"]
        }).json()["id"]
        response = client.post(
            f"/api/tasks/{task_id}/import",
            files={"excel": (sample.name, excel, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert response.status_code == 200
        assert response.json()["imported"] == 20
        item = client.get(f"/api/tasks/{task_id}/items", params={"page_size": 100}).json()["items"][0]
        assert item["interaction_count"] == item["like_count"] + item["comment_count"] + item["share_count"]
        assert item["view_count"] > 0


def test_downloaded_import_template_uses_split_metrics():
    with TestClient(app) as client:
        response = client.get("/api/import-template")
        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        headers = [cell.value for cell in workbook.active[1]]
        assert headers == [
            "平台", "标题", "发布人", "发布时间", "正文", "源链接",
            "点赞量", "评论量", "转发量", "阅读/播放量", "图片文件", "视频文件",
        ]


def test_task_complete_and_reopen():
    with TestClient(app) as client:
        task_id = client.post("/api/tasks", json={
            "name": "测试任务", "keywords": ["测试"], "platforms": ["微博"]
        }).json()["id"]
        completed = client.post(f"/api/tasks/{task_id}/complete").json()
        assert completed["status"] == "completed"
        assert completed["analysis_enabled"] is False
        reopened = client.post(f"/api/tasks/{task_id}/reopen").json()
        assert reopened["status"] == "running"


def test_report_generation_state_and_idempotency(monkeypatch):
    async def fake_report(*_args):
        with SessionLocal() as db:
            pending = db.query(ReportVersion).order_by(ReportVersion.id.desc()).first()
            assert pending is not None
            assert pending.generation_status == "generating"
            assert pending.content == ""
        return "一、任务概况\n测试报告内容"

    monkeypatch.setattr("app.api.reports.generate_report", fake_report)
    with TestClient(app) as client:
        task_id = client.post("/api/tasks", json={
            "name": "报告状态测试", "keywords": ["测试"], "platforms": ["微博"]
        }).json()["id"]
        client.post(
            f"/api/tasks/{task_id}/import",
            files={"excel": ("data.xlsx", excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        item_id = client.get(f"/api/tasks/{task_id}/items").json()["items"][0]["id"]
        client.put(f"/api/items/{item_id}/analysis", json={
            "sentiment": "neutral", "risk_level": "low", "reason": "测试研判",
            "topics": ["测试"], "change_note": "测试修正",
        })
        client.post(f"/api/tasks/{task_id}/complete")

        assert client.get(f"/api/tasks/{task_id}/reports/status").json()["state"] == "available"
        response = client.post(f"/api/tasks/{task_id}/reports")
        assert response.status_code == 200
        assert response.json()["generation_status"] == "completed"
        status_payload = client.get(f"/api/tasks/{task_id}/reports/status").json()
        assert status_payload["state"] == "ready"
        assert status_payload["report_id"] == response.json()["id"]
        duplicate = client.post(f"/api/tasks/{task_id}/reports")
        assert duplicate.status_code == 409
        assert "无需重复生成" in duplicate.json()["detail"]


def test_strategy_generation_state_and_idempotency(monkeypatch):
    task_id_for_generation = 0
    added_during_generation = False

    async def fake_strategy(*_args):
        nonlocal added_during_generation
        with SessionLocal() as db:
            pending = db.query(StrategyVersion).order_by(StrategyVersion.id.desc()).first()
            assert pending is not None
            assert pending.generation_status == "generating"
            assert pending.content == ""
            pending_id = pending.id
            expected_count = pending.analyzed_count
            # 模拟策略生成期间又有一条数据完成研判。状态接口仍应返回
            # 当前版本锁定的数量，而不是不断变化的全局已研判数量。
            if not added_during_generation:
                from app.services.snapshots import current_evidence as database_evidence
                expanded_evidence = database_evidence(db, task_id_for_generation) + [{
                    "item_id": 999,
                    "analysis_id": 999,
                    "title": "生成期间新增研判",
                    "platform": "微博",
                    "content": "这是策略生成期间完成研判的新数据。",
                    "sentiment": "neutral",
                    "risk_level": "low",
                    "reason": "新增研判",
                }]
                monkeypatch.setattr(
                    "app.api.strategies.current_evidence",
                    lambda _db, _task_id: expanded_evidence,
                )
                added_during_generation = True

        with SessionLocal() as observer_db:
            from app.api.strategies import _strategy_state
            generating = _strategy_state(observer_db, task_id_for_generation)
            assert generating["state"] == "generating"
            assert generating["analyzed_count"] == expected_count
            assert generating["strategy_id"] == pending_id
        return "一、态势判断\n测试应对策略"

    monkeypatch.setattr("app.api.strategies.generate_strategy", fake_strategy)
    with TestClient(app) as client:
        task_id = client.post("/api/tasks", json={
            "name": "策略状态测试", "keywords": ["测试"], "platforms": ["微博"]
        }).json()["id"]
        task_id_for_generation = task_id
        client.post(
            f"/api/tasks/{task_id}/import",
            files={"excel": ("data.xlsx", excel_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        item_id = client.get(f"/api/tasks/{task_id}/items").json()["items"][0]["id"]
        client.put(f"/api/items/{item_id}/analysis", json={
            "sentiment": "negative", "risk_level": "medium", "reason": "测试研判",
            "topics": ["测试"], "change_note": "测试修正",
        })

        assert client.get(f"/api/tasks/{task_id}/strategies/status").json()["state"] == "available"
        response = client.post(f"/api/tasks/{task_id}/strategies")
        assert response.status_code == 200
        assert response.json()["generation_status"] == "completed"
        state = client.get(f"/api/tasks/{task_id}/strategies/status").json()
        assert state["state"] == "available"
        assert state["analyzed_count"] == 2
        second = client.post(f"/api/tasks/{task_id}/strategies")
        assert second.status_code == 200
        assert second.json()["analyzed_count"] == 2
        ready = client.get(f"/api/tasks/{task_id}/strategies/status").json()
        assert ready["state"] == "ready"
        assert ready["strategy_id"] == second.json()["id"]
        duplicate = client.post(f"/api/tasks/{task_id}/strategies")
        assert duplicate.status_code == 409
